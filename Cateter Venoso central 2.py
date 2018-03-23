import csv
import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook

root = tk.Tk()
root.withdraw()

txt_file_path = filedialog.askopenfilename(filetypes=(('Text files', 'txt'),))

with open(txt_file_path) as f:
    content = f.readlines()
content = [x.strip('\n') for x in content] 
for line in content:
    print (line)

rows = (
    (88, 46, 57),
    (89, 38, 12),
    (23, 59, 78),
    (56, 21, 98),
    (24, 18, 43),
    (34, 15, 67)
)

excel_file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")]) 
wb = load_workbook(excel_file_path)
ws = wb['Log videos CVC']
ws['A1'] = 'hola'
csv.register_dialect('myDialect', delimiter = ' ')
with open(txt_file_path, 'r') as f:
    reader = csv.reader(f, dialect='myDialect')
    for row in reader:
        print(row)
        ws.append(row)
wb.save(excel_file_path)